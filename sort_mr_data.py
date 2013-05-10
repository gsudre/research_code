'''

Given a temporary folder and a destination folder, organizes the files downloaded form the DICOM server and creates new mask ids in the Excel spreadsheet.

Note that this script assumes the data has been collected properly. In the cases when the study is incomplete, and there are more than one scan session per maskID, upload the incomplete files manually to the directory structure!

Gustavo Sudre, 02/2013

'''

from openpyxl.reader.excel import load_workbook
import os
import glob
import tarfile
import shutil

fname = '/Users/sudregp/tmp/MR_records.xlsx'
tmpFolder = '/Users/sudregp/Downloads/'
mrFolder = '/Volumes/neuro/MR_data/'

wb = load_workbook(filename=fname)
ws = wb.worksheets[0]

dicoms = glob.glob(tmpFolder + '/*-DICOM*')
print 'Found these DICOM files in ' + tmpFolder + ':\n'
for file in dicoms:
    print file

raw_input('\nPress any key to start sorting, or Ctrl+C to quit...')

# while we still have subjects to uncompress
for file in dicoms:
    # strip out the components of the filename
    bar = file.split('-')
    subjectName = bar[0].split('/')
    subjectName = subjectName[-1]
    mrn = bar[1]
    scanId = bar[3]

    # search backwards for the subject's mask ID. Here, we assume it's the latest mask ID a subject was assigned, so that we don't have to search based on the date too
    curRow = ws.get_highest_row()
    curMaskId = None
    while curRow > 0 and curMaskId is None:
        if int(mrn) == int(ws.cell('A' + str(curRow)).value):
            curMaskId = int(ws.cell('L' + str(curRow)).value)
            print 'Found mask ID ' + str(curMaskId)
        else:
            curRow -= 1
    if curRow == 0:
        print 'Error: ' + subjectName + ' not found in spreadsheet!'

    subjFolder = mrFolder + '/' + subjectName + '-' + mrn + '/'
    targetFolder = subjFolder + '/' + str(curMaskId)

    # if we already have the MRN in our tree, that's our target directory
    dir = os.path.dirname(targetFolder)
    # create directory if not exist
    if not os.path.exists(dir):
        os.makedirs(dir)

    # Unpack the DICOMs to the tree
    print 'Unpacking ' + file
    tar = tarfile.open(file)
    tar.extractall(targetFolder)
    tar.close()

    # Remove leading zeros from scanID
    scanId = str(int(scanId))

    rtData = glob.glob(tmpFolder + '/E' + scanId + '.*')
    if len(rtData) == 0:
        print 'WARNING: Could not find real-time data!'
        rtCopied = 'N'
    else:
        print 'Unpacking ' + rtData[0]
        tar = tarfile.open(rtData[0])
        tar.extractall(targetFolder)
        tar.close()
        rtCopied = 'Y'

        # checking if we have all the physiological data we need
        ecgData = glob.glob(targetFolder + '/E' + scanId + '/ECG_*')
        respData = glob.glob(targetFolder + '/E' + scanId + '/Resp_*')
        if len(ecgData) != len(respData):
            print '\tDifferent numbers of respiration and ECG data!'
        elif len(ecgData) == 1:
            print '\tOnly found one set of physiological data: check that it is a child!'
        elif len(ecgData) != 5:  # adults should have 4 + 1 physiological files
            print '\tFound unexpected number of physiological files (%d)!' % len(ecgData)

    # moving extracted files inside maskId folder
    folder2move = glob.glob(targetFolder + '/*' + mrn + '/*')
    shutil.move(folder2move[0], targetFolder)
    os.rmdir(targetFolder + '/' + subjectName + '-' + mrn + '/')
