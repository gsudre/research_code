''' Module to return information about subjects from the Excel spreadsheet that Bethany keeps.
'''

# Gustavo Sudre, 04/2013


def open_spreadsheet():
    # make sure to save an updated version of the file without password in the tmp folder first!
    fname = r'/Users/sudregp/tmp/MEG_Full_Data.xlsx'

    from openpyxl.reader.excel import load_workbook
    wb = load_workbook(filename=fname)
    ws = wb.worksheets[0]

    return ws


def open_dx_spreadsheet():
    # make sure to save an updated version of the file without password in the tmp folder first!
    fname = r'/Users/sudregp/Documents/dx.xlsx'

    from openpyxl.reader.excel import load_workbook
    wb = load_workbook(filename=fname)
    ws = wb.worksheets[0]

    return ws


def get_subjects_from_excel():
    ws = open_spreadsheet()

    subjs = {}
    cnt = 1

    # while we still have something written in the first collumn
    while ws.cell('A' + str(cnt)).value is not None:
        # check if we should use the data or not
        if ws.cell('F' + str(cnt)).value == 'Y':
            subjs[str(ws.cell('D' + str(cnt)).value)] = str(ws.cell('L' + str(cnt)).value)
        cnt = cnt + 1

    return subjs


def get_all_subjects():
    ws = open_spreadsheet()

    subjs = {}

    # start from the second row down
    cnt = 2

    # while we still have something written in the first collumn
    while ws.cell('A' + str(cnt)).value is not None:
        # if there is something in the Acquisition code column
        if isinstance(ws.cell('I' + str(cnt)).value, unicode):
            subjs[str(ws.cell('I' + str(cnt)).value)] = str(ws.cell('O' + str(cnt)).value)

        cnt = cnt + 1

    return subjs


def get_adults(adhd):
    ''' Returns a list of subject codes, which can be ADHDs or NVs, depending on the argument to the function '''
    ws = open_spreadsheet()

    subjs = []

    # start from the second row down
    cnt = 2

    if adhd:
        look_for = 'ADHD'
    else:
        look_for = 'NV'

    # while we still have something written in the first collumn
    while ws.cell('A' + str(cnt)).value is not None:
        # if there is something in the Acquisition code column
        if isinstance(ws.cell('I' + str(cnt)).value, unicode):
            subj_code = str(ws.cell('I' + str(cnt)).value)
            adult = (str(ws.cell('R' + str(cnt)).value) == 'Adult')
            dx = str(ws.cell('P' + str(cnt)).value)
            if adult and (dx == look_for):
                subjs.append(subj_code)

        cnt = cnt + 1

    return subjs


def get_MRNs(subjs):
    ''' Returns a list of MRNs for each subject in the list subjs '''
    ws = open_spreadsheet()

    mrns = []

    for s in subjs:
        mrn = -1
        for row_idx in range(ws.get_highest_row()):
            subj_code = ws.cell(row=row_idx, column=8).value
            if s == subj_code:
                mrn = ws.cell(row=row_idx, column=12).value
        mrns.append(mrn)

    return mrns


def get_ages(subjs):
    ''' Returns a list of ages for each subject in the list subjs '''
    ws = open_spreadsheet()

    ages = []

    import numpy as np
    for s in subjs:
        age = np.NAN
        for row_idx in range(ws.get_highest_row()):
            subj_code = ws.cell(row=row_idx, column=8).value
            if s == subj_code:
                age = ws.cell(row=row_idx, column=9).value
        ages.append(age)

    return ages


def get_remitted_subjects():
    ws = open_dx_spreadsheet()

    subjs = []
    cnt = 1

    # while we still have something written in the first collumn
    while ws.cell('A' + str(cnt)).value is not None:
        # check if we should use the data or not
        if ws.cell('E' + str(cnt)).value == 'remitted':
            subjs.append(ws.cell('A' + str(cnt)).value)
        cnt = cnt + 1
    return subjs


def get_affected_subjects():
    ws = open_dx_spreadsheet()

    subjs = []
    cnt = 1

    # while we still have something written in the first collumn
    while ws.cell('A' + str(cnt)).value is not None:
        # check if we should use the data or not
        if ws.cell('E' + str(cnt)).value != 'remitted':
            subjs.append(ws.cell('A' + str(cnt)).value)
        cnt = cnt + 1
    return subjs
