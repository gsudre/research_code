''' Script to summarize all entries in a table and only keep the latest date of each test, per subject.

Gustavo Sudre, 07/2014 '''

import numpy as np
from openpyxl import load_workbook
import datetime as dt


fname = '/Users/sudregp/dan/family study questionaire and iq data expanded2.xlsx'
# this is the last column that is not in the format Title - Test
last_column = 'Last Name - Subjects'
mrn_column = 'Medical Record - MRN - Subjects'

data = []
wb = load_workbook(filename = r'%s'%fname)
ws = wb.active
last_idx = ws.get_highest_column() # last active column in the sheet

# grab all column headers and saved them for the future
column_headers = [ws.cell(row=1,column=i+1).value for i in range(last_idx)]

mrn_idx = column_headers.index(mrn_column)
first_idx = column_headers.index(last_column) + 1

# figure out all possible tests, and which columns correspond to their data
test_names = [' - '.join(title.split(' - ')[1:]) for title in column_headers[3:]]
test_names = np.unique(test_names)
test_data = {}
for test in test_names:
    data_cols = [i+1 for i in range(last_idx) if column_headers[i].find(test)>=0]
    test_data[test] = data_cols

# the process of getting unique test names screws up the column order, so let's re-write it the same way we'll be outputting the data
column_headers = [ws.cell(row=1,column=i+1).value for i in range(first_idx)]
for test in test_names:
    test_headers = [ws.cell(row=1,column=c).value for c in test_data[test]]
    column_headers += test_headers
data.append(column_headers)

# figure out all unique subjects
last_row = ws.get_highest_row()
all_mrns = [ws.cell(row=i+1, column=1).value for i in range(1,last_row)]
mrns = np.unique(all_mrns)

# for each subject, find out which test is the most recent one. We assume that the date field is always in the first data column of the test!
for mrn in mrns:
    # find all rows that belong to this subject
    subj_rows = [i+2 for i, x in enumerate(all_mrns) if x == mrn]

    # start the subject's row with the standard info, same across all the subject's rows
    data_row = [ws.cell(row=subj_rows[0],column=i+1).value for i in range(first_idx)]

    # for each test, find out which row has the latest date
    for test in test_names:
        tmp_dates = [ws.cell(row=r, column=test_data[test][0]).value for r in subj_rows]
        test_dates = []
        for idx, date in enumerate(tmp_dates):
            if date is None:
                test_dates.append(dt.datetime(1900, 1, 1, 0, 0))
            else:
                try:
                    test_dates.append(dt.datetime.strptime(date,'%m/%d/%Y'))
                except ValueError:
                    print 'Column %d, Row %d: %s does not look like date.'%(test_data[test][0],subj_rows[idx], date)
        # at this point we have the dates of all occurrences of test. Find out the latest one.
        recent_test = subj_rows[np.argmax(test_dates)]
        recent_data = [ws.cell(row=recent_test,column=c).value for c in test_data[test]]
        data_row += recent_data

    data.append(data_row)

# write output file
import csv
out_fname = fname[:-4] + 'csv'
fout = open(out_fname, 'w')
wr = csv.writer(fout)
wr.writerows(data)
fout.close()