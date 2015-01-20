''' Script to create a table of one MRN per row, and the columns list all the possible NSBs.

Gustavo Sudre, 07/2014 '''

import numpy as np
from openpyxl import load_workbook
import datetime as dt


fname = '/Users/sudregp/dan/merged_aspdata_nsbs07232014.xlsx'
nsb_column = 'SAMPLE NSB'
mrn_column = 'MRN'

data = []
wb = load_workbook(filename = r'%s'%fname)
ws = wb.active
last_idx = ws.get_highest_column() # last active column in the sheet

# grab all column headers and saved them for the future
column_headers = [ws.cell(row=1,column=i+1).value for i in range(last_idx)]

mrn_idx = column_headers.index(mrn_column)
nsb_idx = column_headers.index(nsb_column)

# figure out all unique subjects
last_row = ws.get_highest_row()
all_mrns = [ws.cell(row=i+1, column=mrn_idx+1).value for i in range(1,last_row)]
mrns = np.unique(all_mrns)

# for each subject, find out all the NSBs the subject has
for mrn in mrns:
    # find all rows that belong to this subject
    subj_rows = [i+2 for i, x in enumerate(all_mrns) if x == mrn]

    subj_nsbs = [ws.cell(row=r, column=nsb_idx+1).value for r in subj_rows]
    subj_nsbs = np.unique(subj_nsbs)

    data.append([mrn] + list(subj_nsbs))

# construct headers based on the maximum number of NSBs a subject had
max_nsb = 0
for nsbs in data:
    subj_nsbs = len(nsbs)-1  # remove mrn entry
    max_nsb = max(max_nsb, subj_nsbs)
column_headers = ['MRN'] + ['Possible_NSB%02d'%i for i in range(max_nsb)]

# write output file
import csv
out_fname = fname[:-5] + '_mrnsAndNsbs.csv'
fout = open(out_fname, 'w')
wr = csv.writer(fout)
wr.writerows([column_headers])
wr.writerows(data)
fout.close()