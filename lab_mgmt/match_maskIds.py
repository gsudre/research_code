from openpyxl.reader.excel import load_workbook
import datetime

# open maskIds spreadsheet
fname = '/Volumes/shaw data/fMRI/maskIds.xlsx'
miwb = load_workbook(filename=fname)
miws = miwb.worksheets[0]

# open MR spreadsheet
fname = '/Users/sudregp/tmp/all.xlsx'
mrwb = load_workbook(filename=fname)
mrws = mrwb.worksheets[0]

# for each row in the MR spreadsheet, find the item in the maskId spreadsheet with matching MRN and date.
cnt = 2
while mrws.cell('A' + str(cnt)).value is not None:
    mr_date = mrws.cell('I' + str(cnt)).value
    mr_mrn = int(mrws.cell('A' + str(cnt)).value)
    mr_maskid = mrws.cell('L' + str(cnt)).value
    # only check for maskId if the person scanned (MPRAGE)
    mr_mprage = mrws.cell('W' + str(cnt)).value
    if mr_mprage == 'Y':
        found = False
        mi_cnt = 2
        while not found and miws.cell('A' + str(mi_cnt)).value is not None:
            # sometimes the date is not in the correct format
            mi_date = miws.cell('C' + str(mi_cnt)).value
            if not isinstance(mi_date, datetime.datetime):
                mi_date = datetime.datetime.strptime(mi_date, '%m/%d/%Y')
            mi_mrn = int(miws.cell('B' + str(mi_cnt)).value)
            mi_maskid = miws.cell('A' + str(mi_cnt)).value
            if mi_date == mr_date and mi_mrn == mr_mrn:
                found = True
            else:
                mi_cnt += 1

        if found:
            # check if its a blank in MR. If so, copy it from MI. If not, make sure they match
            if mr_maskid is None:
                mrws.cell('L' + str(cnt)).value = mi_maskid
            elif int(mr_maskid) != int(mi_maskid):
                print "Mask IDs do not match: row " + str(cnt)
        else:
            # MRN x date combination not found in MI, so we report it to look for it later
            print "Not found in MaskIds.xlx: " + str(cnt)
    elif mrws.cell('W' + str(cnt)).value == 'DID NOT SCAN' or mrws.cell('G' + str(cnt)).value == 'DID NOT SCAN':
        print "Did not scan: row " + str(cnt)
        mrws.cell('L' + str(cnt)).value = -1
    else:
        print "Look at server for data: row " + str(cnt)

    # not sure why this is going on, but let's
    cnt += 1

mrwb.save(fname)
