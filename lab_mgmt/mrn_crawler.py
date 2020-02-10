''' Goes through a whole folder recursevely and checks all the XLS, XLSX, CSV, and TXT files for the MRNs in a list '''

from openpyxl import load_workbook
from xlrd import open_workbook
import sys


mrnListFile = sys.argv[1]
dataDir = sys.argv[2]

# get the list of MRNs
mrns = []
fid = open(mrnListFile, 'r')
for line in fid:
    mrns.append(line.rstrip())
fid.close()

import os
for dirName, subdirList, fileList in os.walk(dataDir):
    for fname in fileList:
        fullName = dirName + '/' + fname
        # For CSV and TXT files, just search each line for an occurrence of each of the MRNs.
        if fname.lower().find('.csv') > 0 or fname.lower().find('.txt') > 0:
            print('Crawling', fullName)
            foundMRNs = []
            # let's ignore files that are too big, because they probably contain data, not metadata
            if os.path.getsize(fullName) > 100000000:
                print('\tToo big')
            else:
                fid = open(fullName, 'r')
                for line in fid:
                    if fname.lower().find('.csv') > 0:
                        entries = line.split(',')
                    else:
                        entries = line.split(' ')
                    for mrn in mrns:
                        if str(mrn) in entries and mrn not in foundMRNs:
                            print('\tFound', mrn)
                            foundMRNs.append(mrn)
                fid.close()
        # The idea for XLS and XLSX is the same, but we need to use different libraries. We just iterate over all sheets, and all cells in the sheet. If we can convert it to an integer, then we check if it's one of our MRNs
        elif fname.lower().find('.xlsx') > 0:
            try:
                wb = load_workbook(filename = fullName)#, use_iterators = True)
                print('Crawling', fullName)
                foundMRNs = []
                sheets = wb.sheetnames
                for sheet in sheets:
                    ws = wb[sheet]
                    for row in ws.iter_rows():
                        for cell in row:
                            try:
                                thisMRN = str(cell.internal_value)
                            except:
                                thisMRN = None
                            for mrn in mrns:
                                if mrn == thisMRN and mrn not in foundMRNs:
                                    print('\tFound', mrn)
                                    foundMRNs.append(mrn)
            except:
                print('ERROR: Could not open', fullName, '(password protected?)')
        elif fname.lower().find('.xls') > 0:
            try:
                book = open_workbook(fullName)
                print('Crawling', fullName)
                foundMRNs = []
                for sheet in book.sheets():
                    for row_index in range(sheet.nrows):
                        for cell in sheet.row(row_index):
                            try:
                                thisMRN = str(cell.value)
                            except:
                                thisMRN = None
                            for mrn in mrns:
                                if mrn == thisMRN and mrn not in foundMRNs:
                                    print('\tFound', mrn)
                                    foundMRNs.append(mrn)
            except:
                print('ERROR: Could not open', fullName)
