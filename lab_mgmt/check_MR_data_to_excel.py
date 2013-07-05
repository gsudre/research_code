''' Puts a Y in every column in excel for which we have data in the server, and also creates the scan data sheet to be imported into labmatrix '''

import openpyxl
import glob
from datetime import datetime


def check_for_data(dtype, date, folder):
    mask_ids = glob.glob(folder + '/*')
    for mask in mask_ids:
        dirs = glob.glob(mask + '/*')
        dir_num = [d for d, path in enumerate(dirs) if (path.find(date.strftime('%Y_%m_%d')) > 0) or (path.find(date.strftime('%Y%m%d')) > 0)]
        if len(dir_num) > 0:
            # we found a directory with the specific date inside the mask id
            for d in dir_num:
                # now, get the name of the text file
                txtfile = glob.glob(dirs[d] + '/README*.txt')
                fid = open(txtfile[0])
                text = fid.read()
                pos = text.lower().find(dtype)
                if pos > 0:
                    return 1

    # if we get to here, we didn't find the type fo data in any of the matching folders
    print('Data {} not found on {} in {}').format(dtype, date.strftime('%Y_%m_%d'), folder)
    return -1


def check_date(date, folder):
    mask_ids = glob.glob(folder + '/*')
    for mask in mask_ids:
        dirs = glob.glob(mask + '/*')
        dir_num = [d for d, path in enumerate(dirs) if (path.find(scan_date.strftime('%Y_%m_%d')) > 0) or (path.find(scan_date.strftime('%Y%m%d')) > 0)]
        if len(dir_num) > 0:
            # we found a directory with the specific date inside the mask id
            return True
    return False


path = '/Volumes/neuro/MR_data/'

has = {}

# open spreadsheet
fname = r'/Users/sudregp/tmp/visit_records.xlsx'

list_mod = ['rage', 'fmri', 'rest', 'edti', 'clinical', 'fat_sat']  # note that these names need to be found in the README file!
mod_in_excel = ['W', 'X', 'Z', 'AA', 'V', 'AC']
scan_headers = ['MRN', 'Type', 'Scanner', 'Mask ID', 'Task / MEG ID', 'Notes', 'MPRAGE Quality']
# rest,stop task,MPRAGE,clinical,T2,eDTI
# 3TA,3TB,1.5,MEG 

from openpyxl.reader.excel import load_workbook
wb = load_workbook(filename=fname)
ws = wb.worksheets[0]

# get a list of the subjects in the server
subjs = glob.glob(path + '/*')

# for each subject in the server
for row_idx in range(2, ws.get_highest_row()):
    mrn = str(ws.cell('A' + str(row_idx)).value)
    last_name = ws.cell('E' + str(row_idx)).value
    first_name = ws.cell('F' + str(row_idx)).value
    scan_date = ws.cell('I' + str(row_idx)).value

    # if names have a - or space, replace by _
    last_name = last_name.replace('-', '_')
    first_name = first_name.replace('-', '_')
    last_name = last_name.replace(' ', '_')
    first_name = first_name.replace(' ', '_')
    last_name = last_name.replace('\'', '_')

    for m, mod in enumerate(list_mod):
        has[mod] = (ws.cell(mod_in_excel[m] + str(row_idx)).value == 'Y')

    # check whether we have a folder for the subject
    dir_num = [l for l, folder in enumerate(subjs)
               if (folder.find(mrn) > 0 and
                   folder.find(last_name.upper()) > 0 and
                   folder.find(first_name.upper()) > 0)]

    # check whether the subject scanned
    scanned = ws.cell('L' + str(row_idx)).value != -1

    if len(dir_num) == 0 and scanned:
        # subject not found in the server. Check if we expected fMRI data
        missing = [mode for mode, val in has.iteritems() if val]
        if len(missing) > 0:
            print('Subject {}, {} ({}) not in the server. On {}, expected: {}').format(last_name, first_name, mrn, scan_date.strftime('%Y_%m_%d'), missing)
    elif scanned:
        # subject is in the server. First, check whether we have that particular scan for the subject

        if check_date(scan_date, subjs[dir_num[0]]):
            # if we do, then look for the types of data listed in the spreadsheet
            for mod in list_mod:
                if has[mod]:
                    check_for_data(mod, scan_date, subjs[dir_num[0]])
        else:
            missing = [mode for mode, val in has.iteritems() if val]
            if len(missing) > 0:
                print('No scans on {} for {}, {} ({}). Expected: {}').format(scan_date.strftime('%Y_%m_%d'), last_name, first_name, mrn, missing)
